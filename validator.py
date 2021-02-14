import pandas as pd
from openpyxl import load_workbook
from mirri.io.writers.error import ErrorLog, Error
from mirri.io.parsers.mirri_excel import _parse_mirri_v20200601
from mirri.settings import ( MARKERS, MIRRI_FIELDS, LOCATIONS, GROWTH_MEDIA, 
                            GENOMIC_INFO, STRAINS, LITERATURE_SHEET, SEXUAL_STATE_SHEET, 
                            RESOURCE_TYPES_VALUEs, FORM_OF_SUPPLY_SHEET,
                            PLOIDY_SHEET, ONTOBIOTOPE, MARKERS)

TYPES_TRANSLATOR = {'object': str, 'datetime64[ns]': 'datetime', 'int64': int, 'float64': float, 'float32': float}
#fhand = sys.argv[1]
fhand = r"C:\Users\jbravo\Desktop\KPD_MIRRI-IS_dataset_v20201116_v1.2.xlsx"
excel = load_workbook(fhand)
strain = pd.read_excel(fhand, 'Strains', index_col=None)
excelDict = strain.to_dict()  


SHEETS = [
    {"name": LOCATIONS, "acronym": "GOD", "columns":[("ID", True), ("Country", True), ("Region", False), ("City", False), ("Locality", True)]}, 
    {"name": GROWTH_MEDIA, "acronym": "GMD", "columns":[("Acronym", True), ("Description", True), ("Full description", False)]}, 
    {"name": GENOMIC_INFO, "acronym": "GID", "columns":[("Strain AN", False), ("Marker", False), ("INSDC AN", False), ("Sequence", False)]},
    {"name": STRAINS, "acronym": "STD", "columns": [(field['label'], field['mandatory']) for field in MIRRI_FIELDS]},
    {"name": LITERATURE_SHEET, "acronym": "LID", "columns":[("ID", True), ("Full reference", True), ("Authors", True), ("Title", True), ("Journal", True), ("Year", True), ("Volume", True), ("Issue", False), ("First page", True), ("Last page", False), ("Book title", False), ("Editors", False), ("Publisher", False)]}, 
    {"name": SEXUAL_STATE_SHEET, "acronym": "SSD", "columns":[]}, 
    {"name": RESOURCE_TYPES_VALUEs, "acronym": "RTD", "columns":[]}, 
    {"name": FORM_OF_SUPPLY_SHEET, "acronym": "FSD", "columns":[]},
    {"name": PLOIDY_SHEET, "acronym": "PLD", "columns":[]},
    {"name": ONTOBIOTOPE, "acronym": "OTD", "columns":[("ID", False), ("Name", False)]}, 
    {"name": MARKERS, "acronym": "MKD", "columns":[("Acronym", False), ("Marker", False)]}
    ]



#validate excel structure
def validate_excel(excel, strain, excelDict):
    sheets = excel.sheetnames
    errors = []
    excel_name = fhand.split('\\')[-1].rstrip('.xlsx')
    error_log = ErrorLog(excel_name)

    #see if all sheets are there
    for sheet in SHEETS:
        if (sheet['name'] in sheets):
            # print('Valid ' + sheet['name']) 
            sheetEx = excel[sheet['name']]
            if len(sheet['columns']) > 0:
                #validate columns of each sheet
                errors.extend(validate_sheet(sheetEx, sheet))
        else:
            if (sheet['name'] not in ["Ploidy", "Forms of supply", "Resource types values"]):
                # print(f'{sheet["name"]} is not present')
                errors.append(Error(f"The '{sheet['name']}' sheet is missing. Please check the provided excel template", sheet['name']))
    # if len(errors) == 0:
    errors.extend(validation_data(strain, excelDict))

    for error in errors:
        error_log.add_error(error)

    print('Writing to file...')
    error_log.write(f'.\logs')
    return error_log
    

#validate columns
def validate_sheet(sheetEx, sheet):
    errors = []
    columns = next(sheetEx.iter_rows(min_row=1, max_row=1))
    headers = [c.value for c in columns]
    print(headers)
    for col in sheet['columns']:
        if col[0] not in headers and col[1]:
            errors.append(Error(f"The '{col[0]}' is a mandatory field. The column can not be empty.", col[0]))
    return errors
    

def checkTypes(strain, MIRRI_FIELDS, excelDict):
    # Find the columns where each value is null
    stra = strain.dropna(how='all', axis=1)
    types1 = stra.dtypes
    error = []
    types2 = {}
    
    try:
        stra["Recommended growth temperature"] = pd.to_numeric(stra["Recommended growth temperature"], errors='coerce')
    except ValueError:
        # print("string cannot be float")
        error.append(Error("The 'Recommended growth temperature' column has an invalide data type."))
    
    for col, type1 in zip(types1.index, types1):
        if type1.name not in list(TYPES_TRANSLATOR.keys()):
            error.append(Error(f'The "{col}" column has an invalide data type.'))
        types2[col] = TYPES_TRANSLATOR[type1.name]

    for field in MIRRI_FIELDS: 
        if field['label'] in types2:
            if types2[field['label']] != field['type']:
                error.append(Error(f'The "{field["label"]}" column has an invalide data type.'))

def validation_data(strain, excelDict):
    required = [field['label'] for field in MIRRI_FIELDS if field['mandatory']]
    errors = []

    for _, row in strain.iterrows():
        for col, value in row.items():
            #verify where the value is nan and required
            if str(value) == 'nan' and col in required:
                errors.append(Error(f"The '{col}' is missing for strain with Accession Number {row['Accession number']}", row['Accession number']))
                   
    checkTypes(strain, MIRRI_FIELDS, excelDict)

    parsed_excel = _parse_mirri_v20200601(fhand, False)
    for _strain, _errors in parsed_excel['errors'].items():
        for error in _errors:
            errors.append(Error(error['message'].strip('\"\''), _strain))
    return errors
    

#validation(strain, excelDict)
validate_excel(excel, strain, excelDict)