import csv
from copy import deepcopy
from openpyxl.workbook.workbook import Workbook


from mirri import rgetattr
from mirri.settings import GROWTH_MEDIA, MIRRI_FIELDS, DATA_DIR, PUBLICATION_FIELDS
from mirri.io.parsers.mirri_excel import NAGOYA_TRANSLATOR, RESTRICTION_USE_TRANSLATOR

INITIAL_SEXUAL_STATES = [
    "Mata",
    "Matalpha",
    "Mata/Matalpha",
    "Mata",
    "Matb",
    "Mata/Matb",
    "MTLa",
    "MTLalpha",
    "MTLa/MTLalpha",
    "MAT1-1",
    "MAT1-2",
    "MAT1",
    "MAT2",
    "MT+",
    "MT-",
    "MT+",
    "MT-",
    "H+",
    "H-",
]
MARKER_FIELDS = [
    {"attribute": "acronym", "label": "Acronym", "mandatory": True},
    {"attribute": "marker", "label": "Marker", "mandatory": True},
]
MARKER_DATA = [
    {"acronym": "16S rRNA", "marker": "16S rRNA"},
    {"acronym": "ACT", "marker": "Actin"},
    {"acronym": "CaM", "marker": "Calmodulin"},
    {"acronym": "EF-1α", "marker": "elongation factor 1-alpha (EF-1α)"},
    {"acronym": "ITS", "marker": "nuclear ribosomal Internal Transcribed Spacer (ITS)"},
    {"acronym": "LSU", "marker": "nuclear ribosomal Large SubUnit (LSU)"},
    {"acronym": "RPB1", "marker": "Ribosomal RNA-coding genes RPB1"},
    {"acronym": "RPB2", "marker": "Ribosomal RNA-coding genes RPB2"},
    {"acronym": "TUBB", "marker": "β-Tubulin"},
]

REV_RESTRICTION_USE_TRANSLATOR = {v: k for k, v in RESTRICTION_USE_TRANSLATOR.items()}
REV_NAGOYA_TRANSLATOR = {v: k for k, v in NAGOYA_TRANSLATOR.items()}
PUB_HEADERS = [pb["label"] for pb in PUBLICATION_FIELDS]


def write_mirri_excel(path, strains, growth_media, version):
    if version == "20200601":
        _write_mirri_excel_20200601(path, strains, growth_media)


def _write_mirri_excel_20200601(path, strains, growth_media):
    wb = Workbook()

    write_markers_sheet(wb)

    ontobiotope_path = DATA_DIR / "ontobiotopes.csv"
    write_ontobiotopes(wb, ontobiotope_path)

    write_growth_media(wb, growth_media)
    growth_media_indexes = [str(gm.acronym) for gm in growth_media]

    locations = {}
    publications = {}
    sexual_states = set(deepcopy(INITIAL_SEXUAL_STATES))
    genomic_markers = {}
    strains_data = _deserialize_strains(strains, locations, growth_media_indexes,
                                        publications, sexual_states, genomic_markers)
    strains_data = list(strains_data)

    # write strain to generate indexed data
    strain_sheet = wb.create_sheet("Strains")
    strain_sheet.append([field["label"] for field in MIRRI_FIELDS])
    for strain_row in strains_data:
        strain_sheet.append(strain_row)
    redimension_cell_width(strain_sheet)

    # write locations
    loc_sheet = wb.create_sheet("Geographic origin")
    loc_sheet.append(["ID", "Country", "Region", "City", "Locality"])
    for index, loc_index in enumerate(locations.keys()):
        location = locations[loc_index]
        row = [index, location.country, location.state, location.municipality,
               loc_index]
        loc_sheet.append(row)
    redimension_cell_width(loc_sheet)

    # write publications
    pub_sheet = wb.create_sheet("Literature")
    pub_sheet.append(PUB_HEADERS)
    for publication in publications.values():
        row = []
        for pub_field in PUBLICATION_FIELDS:
            # if pub_field['attribute'] == 'id':
            #     value = index
            value = getattr(publication, pub_field['attribute'], None)
            row.append(value)
        pub_sheet.append(row)
    redimension_cell_width(pub_sheet)

    # write sexual states
    sex_sheet = wb.create_sheet("Sexual states")
    for sex_state in sorted(list(sexual_states)):
        sex_sheet.append([sex_state])
    redimension_cell_width(sex_sheet)

    # write genetic markers
    markers_sheet = wb.create_sheet("Genomic information")
    markers_sheet.append(['Strain AN', 'Marker', 'INSDC AN', 'Sequence'])
    for strain_id, markers in genomic_markers.items():
        for marker in markers:
            row = [strain_id, marker.marker_type, marker.marker_id, marker.marker_seq]
            markers_sheet.append(row)
    redimension_cell_width(markers_sheet)

    del wb["Sheet"]
    wb.save(str(path))


def _deserialize_strains(strains, locations, growth_media_indexes,
                         publications, sexual_states, genomic_markers):
    for strain in strains:
        strain_row = []
        for field in MIRRI_FIELDS:
            attribute = field["attribute"]

            if attribute == "id":
                value = strain.id.strain_id
            elif attribute == "restriction_on_use":
                value = rgetattr(strain, attribute)
                if value is not None:
                    value = REV_RESTRICTION_USE_TRANSLATOR[value]
            elif attribute == "nagoya_protocol":
                value = rgetattr(strain, attribute)
                if value:
                    value = REV_NAGOYA_TRANSLATOR[value]
            elif attribute == "other_numbers":
                value = rgetattr(strain, attribute)
                if value is not None:
                    value = [f"{on.collection} {on.number}" for on in value]
                    value = "; ".join(value)
            elif attribute == 'other_denominations':
                od = strain.other_denominations
                value = "; ".join(od) if od else None
            elif attribute in (
                "is_from_registered_collection",
                "is_subject_to_quarantine",
                "is_potentially_harmful",
                "genetics.gmo",
                "taxonomy.interspecific_hybrid"
            ):
                value = rgetattr(strain, attribute)
                if value is True:
                    value = 2
                elif value is False:
                    value = 1
                else:
                    value = None
            elif attribute == "taxonomy.taxon_name":
                value = strain.taxonomy.long_name
            elif attribute in ("deposit.date", "collect.date", "isolation.date",
                               'catalog_inclusion_date'):
                value = rgetattr(strain, attribute)
                value = value.strfdate if value else None
            elif attribute == "growth.recommended_media":
                value = rgetattr(strain, attribute)
                if value is not None:
                    for gm in value:
                        gm = str(gm)
                        if gm not in growth_media_indexes:
                            print(gm, growth_media_indexes)
                            msg = f"Growth media {gm} not in the provided ones"
                            continue
                            raise ValueError(msg)
                    value = "/".join(value)
            elif attribute in ('growth.tested_temp_range',
                               "growth.recommended_temp"):
                value = rgetattr(strain, attribute)
                if value:
                    value = f'{value["min"]}; {value["max"]}'
            elif attribute == "form_of_supply":
                value = rgetattr(strain, attribute)
                value = ";".join(value)
            elif attribute == "collect.location.coords":
                lat = strain.collect.location.latitude
                long = strain.collect.location.longitude
                if lat is not None and long is not None:
                    value = f"{lat};{long}"
                else:
                    value = None

            elif attribute == "collect.location":
                location = strain.collect.location
                loc_index = _build_location_index(location)
                if loc_index is None:
                    continue
                if loc_index not in locations:
                    locations[loc_index] = location
                value = loc_index
            elif attribute in ("abs_related_files", "mta_files"):
                value = rgetattr(strain, attribute)
                value = ";".join(value) if value else None
            elif attribute == "taxonomy.organism_type":
                value = rgetattr(strain, attribute)
                if value:
                    value = "; ".join([str(v.code) for v in value])

            elif attribute == "history":
                value = rgetattr(strain, attribute)
                if value is not None:
                    value = " < ".join(value)
            elif attribute == "genetics.sexual_state":
                value = rgetattr(strain, attribute)
                if value:
                    sexual_states.add(value)
            elif attribute == "genetics.ploidy":
                value = rgetattr(strain, attribute)
            elif attribute == "taxonomy.organism_type":
                organism_types = rgetattr(strain, attribute)
                if organism_types is not None:
                    value = [org_type.code for org_type in organism_types]
                    value = ";".join(value)
            elif attribute == 'publications':
                value = []
                for pub in strain.publications:
                    value.append(pub.id)
                    if pub.id not in publications:
                        publications[pub.id] = pub
                value = ';'.join(str(v) for v in value) if value else None
            elif attribute == 'genetics.plasmids':
                value = rgetattr(strain, attribute)
                if value is not None:
                    value = ';'.join(value)
            else:
                value = rgetattr(strain, attribute)

            strain_row.append(value)
        genomic_markers[strain.id.strain_id] = strain.genetics.markers
        yield strain_row


def _build_location_index(location):
    index = []
    if location.country:
        index.append(location.country)
    if location.site:
        index.append(location.site)
    return ';'.join(index) if index else None


def write_markers_sheet(wb):
    sheet = wb.create_sheet("Markers")
    _write_work_sheet(
        sheet,
        labels=[f["label"] for f in MARKER_FIELDS],
        attributes=[f["attribute"] for f in MARKER_FIELDS],
        data=MARKER_DATA,
    )
    redimension_cell_width(sheet)


def write_ontobiotopes(workbook, ontobiotype_path):
    ws = workbook.create_sheet("Ontobiotope")
    with ontobiotype_path.open() as fhand:
        for row in csv.reader(fhand, delimiter="\t"):
            ws.append(row)
    redimension_cell_width(ws)


def _write_work_sheet(sheet, labels, attributes, data):
    sheet.append(labels)
    for row in data:
        row_data = [row[field] for field in attributes]
        sheet.append(row_data)

    redimension_cell_width(sheet)


def write_growth_media(wb, growth_media):
    ws = wb.create_sheet(GROWTH_MEDIA)
    ws.append(["Acronym", "Description", "Full description"])
    for growth_medium in growth_media:
        row = [
            growth_medium.acronym,
            growth_medium.description,
            growth_medium.full_description,
        ]
        ws.append(row)
    redimension_cell_width(ws)


def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                max_ = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
