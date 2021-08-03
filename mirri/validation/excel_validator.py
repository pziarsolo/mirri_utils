import re
from pathlib import Path
from io import BytesIO
from zipfile import BadZipfile
from datetime import datetime
from calendar import monthrange

from openpyxl import load_workbook

from mirri.io.parsers.excel import workbook_sheet_reader, get_all_cell_data_from_sheet
from mirri.validation.error_logging import ErrorLog, Error
from mirri.validation.tags import (CHOICES, COLUMNS, COORDINATES, CROSSREF, CROSSREF_NAME, DATE,
                                   ERROR_CODE, FIELD, MANDATORY, MATCH,
                                   MISSING, MULTIPLE, NAGOYA, NUMBER, REGEXP, ROW_VALIDATION, SEPARATOR, TAXON,
                                   TYPE, UNIQUE, VALIDATION, VALUES, BIBLIO)
from mirri.settings import LOCATIONS, SUBTAXAS
from mirri.validation.validation_conf_20200601 import MIRRI_20200601_VALLIDATION_CONF


def validate_mirri_excel(fhand, version="20200601"):
    if version == "20200601":
        configuration = MIRRI_20200601_VALLIDATION_CONF
    else:
        raise NotImplementedError("Only version20200601 is implemented")

    return validate_excel(fhand, configuration)


def validate_excel(fhand, configuration):
    validation_conf = configuration['sheet_schema']
    cross_ref_conf = configuration['cross_ref_conf']
    in_memory_sheet_conf = configuration['keep_sheets_in_memory']
    excel_name = Path(fhand.name).stem
    error_log = ErrorLog(excel_name)

    try:
        workbook = load_workbook(filename=BytesIO(
            fhand.read()), read_only=True, data_only=True)
    except (BadZipfile, IOError):
        error = Error('EXL00', fhand.name, fhand.name)
        error_log.add_error(error)
        return error_log

    # excel structure errors
    structure_errors = list(validate_excel_structure(workbook, validation_conf))
    if structure_errors:
        for error in structure_errors:
            error = Error(error[ERROR_CODE], pk=error['id'],
                          data=error['value'])
            error_log.add_error(error)

        return error_log

    crossrefs = get_all_crossrefs(workbook, cross_ref_conf)
    in_memory_sheets = get_all_in_memory_sheet(workbook, in_memory_sheet_conf)
    content_errors = validate_content(workbook, validation_conf,
                                      crossrefs, in_memory_sheets)

    for error in content_errors:
        # if error[ERROR_CODE] == 'STD43':
        #     continue
        error = Error(error[ERROR_CODE], pk=error['id'], data=error['value'])

        error_log.add_error(error)
    return error_log


def validate_excel_structure(workbook, validation_conf):
    for sheet_name, sheet_conf in validation_conf.items():
        mandatory = sheet_conf.get(VALIDATION, {}).get(TYPE, None)
        mandatory = mandatory == MANDATORY

        error_code = sheet_conf.get(VALIDATION, {}).get(ERROR_CODE, False)
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            sheet = None

        if sheet is None:
            if mandatory:
                yield {'id': None, 'sheet': sheet_name, 'field': None,
                       'error_code': error_code, 'value': None}
            continue

        headers = _get_sheet_headers(sheet)
        for column in sheet_conf.get(COLUMNS):
            field = column[FIELD]
            for step in column.get(VALIDATION, []):
                if step[TYPE] == MANDATORY and field not in headers:
                    yield {'id': None, 'sheet': sheet_name, 'field': field,
                           'error_code': step[ERROR_CODE], 'value': None}


def _get_sheet_headers(sheet):
    first_row = next(sheet.iter_rows(min_row=1, max_row=1))
    return [c.value for c in first_row]


def _get_values_from_columns(workbook, sheet_name, columns):
    indexed_values = {}
    for row in workbook_sheet_reader(workbook, sheet_name):
        for col in columns:
            indexed_values[str(row.get(col))] = ""

    return indexed_values


def get_all_crossrefs(workbook, cross_refs_names):
    crossrefs = {}
    for ref_name, columns in cross_refs_names.items():
        if columns:
            crossrefs[ref_name] = _get_values_from_columns(workbook, ref_name,
                                                               columns)
        else:
            try:
                crossrefs[ref_name] = get_all_cell_data_from_sheet(workbook, ref_name)
            except ValueError as error:
                if 'sheet is missing' in str(error):
                    crossrefs[ref_name] = []
                else:
                    raise

    return crossrefs


def get_all_in_memory_sheet(workbook, in_memory_sheet_conf):
    in_memory_sheets = {}
    for sheet_conf in in_memory_sheet_conf:
        sheet_name = sheet_conf['sheet_name']
        indexed_by = sheet_conf['indexed_by']
        rows = workbook_sheet_reader(workbook, sheet_name)
        indexed_rows = {row[indexed_by]: row for row in rows}
        in_memory_sheets[sheet_name] = indexed_rows

    return in_memory_sheets


def validate_content(workbook, validation_conf, crossrefs, in_memory_sheets):
    for sheet_name in validation_conf.keys():
        sheet_conf = validation_conf[sheet_name]
        sheet_id_column = sheet_conf['id_field']
        shown_values = {}
        row_validation_steps = sheet_conf.get(ROW_VALIDATION, None)
        for row in workbook_sheet_reader(workbook, sheet_name):
            id_ = row.get(sheet_id_column, None)
            if id_ is None:
                error_code = _get_missing_row_id_error(sheet_id_column,
                                                       sheet_conf)
                yield {'id': id_, 'sheet': sheet_name,
                       'field': sheet_id_column,
                       'error_code': error_code, 'value': None}
                continue
            do_have_cell_error = False
            for column in sheet_conf[COLUMNS]:
                label = column[FIELD]
                validation_steps = column.get(VALIDATION, None)
                value = row.get(label, None)
                if validation_steps:
                    error_code = validate_cell(value, validation_steps,
                                               crossrefs, shown_values, label)
                    if error_code is not None:
                        do_have_cell_error = True
                        yield {'id': id_, 'sheet': sheet_name, 'field': label,
                               'error_code': error_code, 'value': value}

            if not do_have_cell_error and row_validation_steps:
                error_code = validate_row(
                    row, row_validation_steps, in_memory_sheets)
                if error_code is not None:
                    yield {'id': id_, 'sheet': sheet_name, 'field': 'row',
                           'error_code': error_code, 'value': 'row'}


def _get_missing_row_id_error(sheet_id_column, sheet_conf):
    error_code = None
    for column in sheet_conf[COLUMNS]:
        if column[FIELD] == sheet_id_column:
            error_code = [step[ERROR_CODE]
                          for step in column[VALIDATION] if step[TYPE] == MISSING][0]
    return error_code


def validate_row(row, validation_steps, in_memory_sheets):
    for validation_step in validation_steps:
        kind = validation_step[TYPE]
        error_code = validation_step[ERROR_CODE]
        if kind == NAGOYA:
            if not is_valid_nagoya(row, in_memory_sheets):
                return error_code
        elif kind == BIBLIO:
            if not is_valid_pub(row):
                return error_code
        else:
            msg = f'{kind} is not a recognized row validation type method'
            raise NotImplementedError(msg)


def validate_cell(value, validation_steps, crossrefs, shown_values, label):

    for step_conf in validation_steps:
        if step_conf[TYPE] == MANDATORY:
            continue
        step_conf['crossrefs_pointer'] = crossrefs
        step_conf['shown_values'] = shown_values
        step_conf['label'] = label
        error_code = validate_value(value, step_conf)

        if error_code is not None:
            return error_code


def is_valid_pub(row):
    title = row.get('Title', None)
    full_reference = row.get('Full reference', None)
    authors = row.get('Authors', None)
    journal = row.get('Journal', None)
    year = row.get('Year', None)
    volumen = row.get('Volumen', None)
    first_page = row.get('First page', None)
    book_title = row.get('Book title', None)
    editors = row.get('Editors', None)
    publishers = row.get('Publishers', None)

    if full_reference:
        return True
    is_journal = bool(title)

    if (is_journal and (not authors  or not journal or not not year or
                        not volumen or not first_page)):
        return False
    if (not is_journal and (not authors or not year or
                            not editors or not publishers or not book_title)):
        return False

    return True


def is_valid_nagoya(row, in_memory_sheets):  # sourcery skip: return-identity
    location_index = row.get('Geographic origin', None)
    if location_index is None:
        country = None
    else:
        geo_origin = in_memory_sheets[LOCATIONS].get(location_index, {})
        country = geo_origin.get('Country', None)

    _date = row.get("Date of collection", None)
    if _date is None:
        _date = row.get("Date of isolation", None)
    if _date is None:
        _date = row.get("Date of deposit", None)
    if _date is None:
        _date = row.get("Date of inclusion in the catalogue", None)
    if _date is not None:
        year = _date.year if isinstance(_date, datetime) else int(str(_date)[:4])
    else:
        year = None

    if year is not None and year >= 2014 and country is None:
        return False

    return True


def is_valid_regex(value, validation_conf):
    if value is None:
        return True
    value = str(value)
    regexp = validation_conf[MATCH]
    multiple = validation_conf.get(MULTIPLE, False)
    separator = validation_conf.get(SEPARATOR, None)

    values = [v.strip() for v in value.split(
        separator)] if multiple else [value]

    for value in values:
        matches_regexp = re.fullmatch(regexp, value)
        if not matches_regexp:
            return False
    return True


def is_valid_crossrefs(value, validation_conf):
    crossref_name = validation_conf[CROSSREF_NAME]
    crossrefs = validation_conf['crossrefs_pointer']
    choices = crossrefs[crossref_name]
    if value is None or not choices:
        return True
    value = str(value)

    multiple = validation_conf.get(MULTIPLE, False)
    separator = validation_conf.get(SEPARATOR, None)
    if value is None:
        return True
    if multiple:
        values = [v.strip() for v in value.split(separator)]
    else:
        values = [value.strip()]

    return all(value in choices for value in values)


def is_valid_choices(value, validation_conf):
    if value is None:
        return True
    choices = validation_conf[VALUES]
    multiple = validation_conf.get(MULTIPLE, False)
    separator = validation_conf.get(SEPARATOR, None)

    if multiple:
        values = [v.strip() for v in str(value).split(separator)]
    else:
        values = [str(value).strip()]

    return all(value in choices for value in values)


def is_valid_date(value, validation_conf):
    if value is None:
        return True
    if isinstance(value, datetime):
        year = value.year
        month = value.month
        day = value.day
    elif isinstance(value, int):
        year = value
        month = None
        day = None
    elif isinstance(value, str):
        value = value.replace('-', '')
        value = value.replace('/', '')
        month = None
        day = None
        try:
            year = int(value[: 4])
            if len(value) >= 6:
                month = int(value[4: 6])
                if len(value) >= 8:
                    day = int(value[6: 8])

        except (IndexError, TypeError, ValueError):
            return False
    else:
        return False

    if year < 1700 or year > datetime.now().year:
        return False
    if month is not None:
        if month < 1 or month > 13:
            return False
        if day is not None and (day < 1 or day > monthrange(year, month)[1]):
            return False
    return True


def is_valid_coords(value, validation_conf=None):
    # sourcery skip: return-identity
    if value is None:
        return True
    try:
        items = [i.strip() for i in value.split(";")]
        latitude = float(items[0])
        longitude = float(items[1])
        if len(items) > 2:
            precision = float(items[2])
        if latitude < -90 or latitude > 90:
            return False
        if longitude < -180 or longitude > 180:
            return False
        return True
    except:
        return False


def is_valid_missing(value, validation_conf=None):
    return value is not None


def is_valid_number(value, validation_conf):
    if value is None:
        return True
    try:
        value = float(value)
    except TypeError:
        return False
    except ValueError:
        return False

    _max = validation_conf.get('max', None)
    _min = validation_conf.get('min', None)
    if (_max is not None and value > _max) or (_min is not None and value < _min):
        return False

    return True


def is_valid_taxon(value, validation_conf=None):
    multiple = validation_conf.get(MULTIPLE, False)
    separator = validation_conf.get(SEPARATOR, ';')

    value = value.split(separator) if multiple else [value]
    for taxon in value:
        taxon = taxon.strip()
        if not _is_valid_taxon(taxon):
            return False
    return True


def _is_valid_taxon(value):
    value = value.strip()
    if not value:
        return True

    items = re.split(r" +", value)
    genus = items[0]

    if len(items) > 1:
        species = items[1]
        if species in ("sp", "spp", ".sp", "sp."):
            return False

        if len(items) > 2:
            for index in range(0, len(items[2:]), 2):
                rank = SUBTAXAS.get(items[index + 2], None)
                if rank is None:
                    print(value)
                    return False

    return True


def is_valid_unique(value, validation_conf):
    label = validation_conf['label']
    shown_values = validation_conf['shown_values']
    if label not in shown_values:
        shown_values[label] = {}

    already_in_file = shown_values[label]
    if value in already_in_file:
        return False

    # NOTE: what's the use of this?
    # What is the expected format for value and shown_values?
    shown_values[label][value] = None

    return True


def is_valid_file(path):
    try:
        with path.open("rb") as fhand:
            error_log = validate_mirri_excel(fhand)
            if "EXL" in error_log.get_errors():
                return False
    except:
        return False

    return True


VALIDATION_FUNCTIONS = {
    MISSING: is_valid_missing,
    REGEXP: is_valid_regex,
    CHOICES: is_valid_choices,
    CROSSREF: is_valid_crossrefs,
    DATE: is_valid_date,
    COORDINATES: is_valid_coords,
    NUMBER: is_valid_number,
    TAXON: is_valid_taxon,
    UNIQUE: is_valid_unique}


def validate_value(value, step_conf):
    kind = step_conf[TYPE]
    try:
        is_value_valid = VALIDATION_FUNCTIONS[kind]
    except KeyError:
        msg = f'This validation type {kind} is not implemented'
        raise NotImplementedError(msg)

    error_code = step_conf[ERROR_CODE]

    if not is_value_valid(value, step_conf):
        return error_code
